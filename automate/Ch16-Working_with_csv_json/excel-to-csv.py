import os
import openpyxl
import csv
from openpyxl.utils import get_column_letter
os.chdir('Ch16-Working_with_csv_json')


def excelToCSV(folder):
    for file in os.listdir(folder):
        if not file.endswith('.xlsx'):
            continue
        wb = openpyxl.load_workbook(file)
        sheet = wb[wb.sheetnames[0]]
        csvFile = os.path.splitext(file)[0] + '_' + wb.sheetnames[0]
        print(f'coverting {file}')
        outputFile = open(f'{csvFile}.csv', 'w', newline='')
        outputWriter = csv.writer(outputFile)
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []
            for colNum in range(1, sheet.max_column + 1):
                rowData.append(
                    sheet[get_column_letter(colNum)+str(rowNum)].value)
            outputWriter.writerow(rowData)
        outputFile.close()
        print(f'{file} to {csvFile}.csv')


if __name__ == "__main__":
    excelToCSV('.')
