#! python3
# blankRowInserter.py - Inserts n blank rows at point m from mySpreadsheet.xlsx.

import openpyxl
import sys

from openpyxl.utils.cell import get_column_letter

if len(sys.argv) > 1:

    try:
        m = int(sys.argv[1])
        n = int(sys.argv[2])
        location = str(sys.argv[3])

    except Exception as e:
        print(e)

    wb = openpyxl.load_workbook(location)
    sheet = wb.active

    # Get coordinates of spreadsheet origin after insertion.
    x1 = get_column_letter(sheet.min_column)
    y1 = str(m)

    # Get coordinates of spreadsheet max.
    x2 = get_column_letter(sheet.max_column)
    y2 = str(sheet.max_row)

    c1 = x1 + y1
    c2 = x2 + y2

    # Translate will update the formulae.
    sheet.move_range(c1 + ":" + c2, rows=n, translate=True)

    wb.save(location)

    print("Operation successful. All rows have been inserted.")

else:
    print("You must structure your argument like so: python blankRowInserter.py m n mySpreadsheet.xlsx")
