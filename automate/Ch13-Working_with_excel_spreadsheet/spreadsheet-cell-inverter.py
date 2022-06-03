#! python3
# sheetInverter.py - Transposes the rows and columns in a spreadsheet
# Author: Priyank Jain
# small adjust: max_row, max_column
# sys_exit(-num) to see the post https://stackoverflow.com/questions/44893807/i-want-to-know-what-exactly-sys-exit-1-returns-in-python
#####################################################################
import openpyxl
import sys
import os
if len(sys.argv) > 2:
    print('Usage: python3 sheetInverter.py <worbook-name>')
    sys.exit(-1)
if not os.path.exists(sys.argv[1]):
    print('The specified file {0} does not exist'.format(sys.argv[1]))
    sys.exit(-2)
inputFile = sys.argv[1]
outputFile = '{0}_inverted{1}'.format(*os.path.splitext(inputFile))
if os.path.exists(outputFile):
    print('Destination file {0} already exists'.format(outputFile))
    sys.exit(-3)
inputWb = openpyxl.load_workbook('example.xlsx')
inputSheet = inputWb.active
outputWb = openpyxl.Workbook()
outputSheet = outputWb.active
maxRow = inputSheet.max_row
maxCol = inputSheet.max_column
maxRow
maxCol
for i in range(1, max(maxRow, maxCol)+1):
    for j in range(1, min(maxRow, maxCol)+1):
        outputSheet.cell(row=i, column=j).value = inputSheet.cell(
            row=j, column=i).value
        outputSheet.cell(row=j, column=i).value = inputSheet.cell(
            row=i, column=j).value
outputWb.save('example1.xlsx')
inputWb.save(inputFile)
print('File inverted successfully')
