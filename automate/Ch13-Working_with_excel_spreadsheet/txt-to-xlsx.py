import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
os.chdir('.\Ch13-Working_with_excel_spreadsheet')
# get files end with txt
toXlsx = []
for x, y, z in os.walk('.'):
    for filenames in z:
        if filenames.endswith('txt'):
            toXlsx.append(filenames)
# write txt into xlsx
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(len(toXlsx)):
    file = open(toXlsx[i], 'r')
    text = file.readlines()
    for j in range(len(text)):
        sheet[get_column_letter(i+1) + str(j+1)].value = text[j]
    sheet.column_dimensions[get_column_letter(
        i+2)].width = len(max(text, key=len))
    file.close()
# bonus
sheet.move_range('A1:B100', rows=1, cols=1)
for i in range(100):
    # can add \n in each row for the later traceback from xlsx.
    sheet['A' + str(i+2)].value = 'Top' + str(i+1)
fontObj = Font(bold=True)
sheet['A1'].value = 'Rank'
sheet['A1'].font = fontObj
sheet['B1'].value = 'Singers'
sheet['B1'].font = fontObj
sheet['C1'].value = 'Song'
sheet['C1'].font = fontObj
wb.save('Billboard100.xlsx')
