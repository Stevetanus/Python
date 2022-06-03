import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.active
fontObj = Font(bold=True)
num = int(input('enter a positive number'))
for i in range(num):
    sheet['A' + str(i+2)] = i+1
    sheet['A' + str(i+2)].font = fontObj
    sheet[get_column_letter(i+2) + '1'] = i+1
    sheet[get_column_letter(i+2) + '1'].font = fontObj
for i in range(num):
    for j in range(num):
        n = (i+1) * (j+1)
        sheet[get_column_letter(j+2) + str(i+2)].value = n
wb.save('practice1.xlsx')
