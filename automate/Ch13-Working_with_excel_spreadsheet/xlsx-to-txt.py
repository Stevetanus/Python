import openpyxl
import os
from openpyxl.utils import get_column_letter
os.chdir('.\Ch13-Working_with_excel_spreadsheet')
# search all xlsx in the working dir.
toTxt = []
for x, y, z in os.walk('.'):
    for filenames in z:
        if filenames.endswith('xlsx'):
            toTxt.append(filenames)
# write each column in xlsx into txt with the file name as the first row's value.
wb = openpyxl.load_workbook(toTxt[0])
sheet = wb.active
for i in range(sheet.max_column):
    count = 0
    for cell in sheet[get_column_letter(i+1)]:
        if not cell.value is None:
            count += 1
    file = open('{}.txt'.format(
        sheet[get_column_letter(i+1) + '1'].value), 'w')
    for text in range(count-1):
        file.write(sheet[get_column_letter(i+1) + str(text+2)].value)
    file.close()
