import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
# 從活頁簿中取得工作表
wb.sheetnames
sheet = wb['Sheet3']
sheet
type(sheet)
sheet.title
anotherSheet = wb.active
anotherSheet
# 從工作表中取得儲存格
sheet = wb['Sheet1']
sheet['A1']
sheet['A1'].value
c = sheet['B1']
c.value
# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value) # 'Row 1, Column 2 is Apples'
'Cell %s is %s' % (c.coordinate, c.value) # 'Cell B1 is Apples'
sheet['c1'].value
sheet.cell(row=1, column=2)
sheet.cell(row=1, column=2).value
for i in range(1, 8, 2): # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)
# 確定工作表的內容大小
sheet.max_row
sheet.max_column
# 欄的字母和數字之間的轉換
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1)
get_column_letter(2)
get_column_letter(27)
get_column_letter(900)
get_column_letter(sheet.max_column)
column_index_from_string('AA')
#  從工作表中取得列和欄
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj)
        print(cellObj.coordinate, cellObj.value)
    print('---- END of ROW ----')
sheet = wb.active
list(sheet.columns)[1]
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
# 寫入Excel檔
wb = openpyxl.Workbook() # Create a blank workbook. W is capital.
wb.sheetnames
sheet = wb.active
sheet.title
sheet.title = 'Spam Bacon Eggs Sheet' # Change title.
wb.sheetnames

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx') # Save the workbook to another xlsx flie.
# create_sheet() and del operator
wb = openpyxl.Workbook()
wb.sheetnames
wb.create_sheet()
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet') # create sheet at index 0.
wb.sheetnames
wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames
del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames
# Writing Values to Cells
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!' # Edit the cell's value
sheet['A1'].value
# 設定儲存格的字型
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font.
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello, world!'
wb.save('styled.xlsx')
wb = openpyxl.Workbook()
sheet = wb['Sheet']
fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'
wb.save('styles.xlsx')
# formula
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')
# row and column dimensions
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')
# merge cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5') # merge these two cells.
sheet['C5'] = 'Two merge cells.'
wb.save('merged.xlsx')
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3') # Split these cells up.
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')
# freeze
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2' # freeze the rows above A2.
wb.save('freezeExample.xlsx')
# charts
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, 1, 1, 1, 10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')
